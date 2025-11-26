# sooratvaziat/views.py
from django.contrib.humanize.templatetags.humanize import intcomma
import jdatetime
from jalali_date.fields import JalaliDateField 
from jalali_date.widgets import AdminJalaliDateWidget 
from datetime import datetime
from datetime import date

from django.utils import timezone

from django.contrib import messages

from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from collections import OrderedDict
from django.views.generic import ListView
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.db.models import Prefetch, Sum, Count, Q
from django.db import transaction
from django.http import HttpResponse, JsonResponse
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
from collections import defaultdict
from datetime import date

#forms 
from django.forms import inlineformset_factory, modelform_factory, HiddenInput, TextInput, Select
from project.forms import ProjectCreateForm, ProjectEditForm
from .forms import MeasurementSessionForm, MeasurementSessionItemForm

#models
from .models import MeasurementSummary, MeasurementSessionItem,DetailedMeasurement,ProjectFinancialSummary, MeasurementSession, MeasurementSessionItem
from project.models import Project, StatusReport
from fehrestbaha.models import DisciplineChoices
from fehrestbaha.models import PriceList, PriceListItem, DisciplineChoices
from accounts.models import ProjectUser

#PDF
from io import BytesIO
from django.template.loader import render_to_string  # برای PDF
from xhtml2pdf import pisa

from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from .mixins import UserProjectMixin

#search
from django.views.decorators.http import require_http_methods
import json
from django.core.paginator import Paginator
#logging
import logging
# utils
from sooratvaziat.utils import (
    gregorian_to_jalali,
    jalali_to_gregorian,
    format_number_int,
    _to_decimal,
    _get_progress_class,
    format_number_decimal,
    get_status_badge,
    format_currency,
    get_project_statistics,
    calculate_financial_metrics,
    get_financial_summary,
    get_recent_events,
    get_project_warnings,
    get_chart_data,
    calculate_project_duration,
    get_last_activity,
    get_project_with_access,
    get_user_project_role,
    can_edit_directly,
    can_view_revisions
)
logger = logging.getLogger(__name__)

@login_required
def riz_metre_financial(request, pk, discipline_choice=None):
    # فقط پروژه‌های کاربر جاری
    project = get_object_or_404(Project, pk=project_id, is_active=True)

    # سوپریوزر و ادمین به همه پروژه‌ها دسترسی دارند
    if user.is_superuser:
        return project
    
    if UserRole.objects.filter(user=user, role='admin', is_active=True).exists():
        return project

    # کاربران عادی
    has_access = (
        project.created_by == user or
        ProjectUser.objects.filter(
            project=project, 
            user=user, 
            is_active=True
        ).exists()
    )
    
    if not has_access:
        raise PermissionDenied("شما دسترسی به این پروژه را ندارید")

    # فیلتر کردن بر اساس پروژه و فهرست بها (اگر مشخص شده باشد)
    qs = MeasurementSessionItem.objects.filter(
        measurement_session_number__project=project,
        is_active=True
    ).select_related('pricelist_item').order_by('pricelist_item__row_number', 'id')
    
    if discipline_choice:
        qs = qs.filter(pricelist_item__price_list__discipline_choice=discipline_choice)
    
    rows = OrderedDict()
    for item in qs:
        pl = item.pricelist_item
        key = getattr(pl, 'row_number', None) or f"_id_{pl.pk}"
        if key not in rows:
            rows[key] = {
                'pricelist_item': pl,
                'row_number': getattr(pl, 'row_number', ''),
                'unit': getattr(pl, 'unit', '') or '',
                'description': getattr(pl, 'row_description', '') or '',
                'total_qty': Decimal('0.00'),
                'unit_price': Decimal('0.00'),
                'line_total': Decimal('0.00'),
            }
        try:
            raw_amount = item.get_total_item_amount()
        except Exception:
            raw_amount = getattr(item, 'total', 0)
        qty = Decimal(str(raw_amount or 0))
        rows[key]['total_qty'] += qty

    # 📘 مرحله بعد: تعیین قیمت و جمع‌ها
    for r in rows.values():
        pl = r['pricelist_item']
        unit_price = None
        for cand in ('price', 'unit_price', 'rate', 'baha'):
            if hasattr(pl, cand):
                val = getattr(pl, cand)
                if val is not None:
                    try:
                        unit_price = Decimal(str(val))
                        break
                    except Exception:
                        unit_price = Decimal('0')
        if unit_price is None:
            unit_price = Decimal('0')

        # ذخیره اعداد به‌صورت Decimal (برای محاسبات) و رشته‌ی فرمت‌شده برای نمایش
        r['unit_price'] = unit_price.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        r['line_total'] = (r['total_qty'] * r['unit_price']).quantize(Decimal('1'), rounding=ROUND_HALF_UP)

        # اضافه کردن فیلدهای نمایش فرمت‌شده:
        r['formatted_total_qty'] = format_number_int(r['total_qty'])
        r['formatted_unit_price'] = format_number_int(r['unit_price'])
        r['formatted_line_total'] = format_number_int(r['line_total'])

    # 📗 حالا شماره‌گذاری فصل‌ها و ردیف‌ها
    chapter_counters = defaultdict(int)
    numbered_rows = []
    prev_chapter = None

    for r in rows.values():
        rn = str(r['row_number'])
        # استخراج فصل: دو کاراکتر اول (اگر کمتر باشه "00")
        chapter = rn[:2] if len(rn) >= 2 else "00"

        # شمارنده فصل را افزایش بده
        chapter_counters[chapter] += 1
        display_number = f"{chapter}-{chapter_counters[chapter]}"  # مثال: 07-1

        r['display_number'] = display_number
        r['chapter'] = chapter

        # فرمت‌های نمایشی برای HTML/CSV/XLSX
        r['formatted_total_qty'] = format_number_int(r['total_qty'])
        r['formatted_unit_price'] = format_number_int(r['unit_price'])
        r['formatted_line_total'] = format_number_int(r['line_total'])

        # مشخص می‌کنیم فصل جدید شروع شده یا نه
        r['is_new_chapter'] = (chapter != prev_chapter)
        prev_chapter = chapter

        numbered_rows.append(r)

    # نام فهرست بها برای نمایش در عنوان
    discipline_label = None
    if discipline_choice:
        discipline_label = dict(DisciplineChoices.choices).get(discipline_choice, 'نامشخص')

    # ----------------- خروجی CSV -----------------
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        filename = f"soorat_mali_project_{project.id}"
        if discipline_choice:
            filename += f"_{discipline_choice}"
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        writer = csv.writer(response)
        writer.writerow(['شماره ردیف', 'شماره آیتم', 'شرح آیتم', 'واحد', 'جمع مقدار', 'قیمت واحد (ریال)', 'جمع ریالی (ریال)'])

        for r in numbered_rows:
            description = getattr(r['pricelist_item'], 'description', '') or r.get('description', '')
            writer.writerow([
                r['display_number'],
                r['row_number'],
                description,
                r['unit'],
                f"{int(r['total_qty']):,}",
                f"{int(r['unit_price']):,}",
                f"{int(r['line_total']):,}",
            ])

        writer.writerow([])
        return response

    # ----------------- خروجی Excel -----------------
    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = f"صورت مالی پروژه {project.project_name}"
        if discipline_label:
            ws.title += f" - {discipline_label}"
        headers = ['شماره ردیف', 'شماره آیتم', 'شرح آیتم', 'واحد', 'جمع مقدار', 'قیمت واحد (ریال)', 'جمع ریالی (ریال)']
        ws.append(headers)

        for r in numbered_rows:
            description = getattr(r['pricelist_item'], 'description', '') or r.get('description', '')
            ws.append([
                r['display_number'],
                r['row_number'],
                description,
                r['unit'],
                int(r['total_qty']),
                int(r['unit_price']),
                int(r['line_total']),
            ])

        # استایل جدول
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                if cell.row == 1:
                    cell.font = Font(bold=True)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"soorat_mali_project_{project.id}"
        if discipline_choice:
            filename += f"_{discipline_choice}"
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response

    # ----------------- خروجی HTML -----------------
    context = {
        'title': f'صورت مالی (ریز مالی) - {project.project_name}',
        'rows': numbered_rows,
        'project': project,
        'discipline_choice': discipline_choice,
        'discipline_label': discipline_label,
    }
    return render(request, 'sooratvaziat/soorat_mali.html', context)

@login_required
def riz_financial_discipline_list(request, pk):
    project = get_project_with_access(request.user, project_pk)
    
    # استخراج رشته‌های منحصر به فرد از آیتم‌های موجود برای پروژه
    disciplines = MeasurementSessionItem.objects.filter(
        measurement_session_number__project=project,
        is_active=True
    ).values_list(
        'pricelist_item__price_list__discipline_choice', 
        flat=True
    ).distinct()

    # تبدیل به لیست از tuples برای استفاده در تمپلیت
    discipline_choices = []
    for discipline in disciplines:
        label = dict(DisciplineChoices.choices).get(discipline, 'نامشخص')
        # محاسبه تعداد آیتم‌ها و جمع مبالغ برای هر فهرست بها
        items = MeasurementSessionItem.objects.filter(
            measurement_session_number__project=project,
            pricelist_item__price_list__discipline_choice=discipline,
            is_active=True
        )
        
        total_amount = Decimal('0')
        for item in items:
            try:
                qty = Decimal(str(item.get_total_item_amount() or 0))
                pl = item.pricelist_item
                unit_price = Decimal('0')
                for cand in ('price', 'unit_price', 'rate', 'baha'):
                    if hasattr(pl, cand):
                        val = getattr(pl, cand)
                        if val is not None:
                            unit_price = Decimal(str(val))
                            break
                total_amount += qty * unit_price
            except:
                continue
        
        discipline_choices.append({
            'value': discipline,
            'label': label,
            'count': items.count(),
            'total_amount': total_amount.quantize(Decimal('1'), rounding=ROUND_HALF_UP),
            'formatted_total_amount': format_number_int(total_amount),
        })

    context = {
        'project': project,
        'disciplines': discipline_choices,
    }
    return render(request, 'sooratvaziat/riz_financial_discipline_list.html', context)

@login_required
def riz_metre_discipline_list(request, pk):
    """
    لیست رشته‌های موجود در پروژه برای ریز متره
    """
    project = get_project_with_access(request.user, pk)
    
    # یافتن رشته‌های موجود در صورت جلسات پروژه
    disciplines_data = MeasurementSession.objects.filter(
        project=project,
        is_active=True,
        items__is_active=True
    ).values(
        'price_list__discipline_choice',
        'price_list__discipline'
    ).annotate(
        item_count=Count('items', filter=Q(items__is_active=True)),
        session_count=Count('id', distinct=True)
    ).filter(item_count__gt=0).order_by('price_list__discipline_choice')
    
    # تبدیل به فرمت مورد نیاز برای تمپلیت
    disciplines = []
    for disc in disciplines_data:

        discipline_value = disc['price_list__discipline_choice']
        discipline_label = dict(DisciplineChoices.choices).get(discipline_value, discipline_value)
        
        disciplines.append({
            'value': discipline_value,
            'label': discipline_label,
            'count': disc['item_count'],
            'session_count': disc['session_count'],
        })
    
    context = {
        'title': f'انتخاب رشته برای ریز متره - {project.project_name}',
        'project': project,
        'disciplines': disciplines,
    }
    return render(request, 'sooratvaziat/riz_metre_discipline_list.html', context)

@login_required
def riz_metre(request, pk, discipline):
    """
    نمایش ریز متره برای پروژه و رشته خاص - نسخه موقت بدون DetailedMeasurement
    """
    project = get_project_with_access(request.user, pk)
    
    # دریافت مستقیم از صورت جلسات
    session_items = MeasurementSessionItem.objects.filter(
        measurement_session_number__project=project,
        measurement_session_number__is_active=True,
        measurement_session_number__price_list__discipline_choice=discipline,
        is_active=True
    ).select_related(
        'pricelist_item',
        'pricelist_item__price_list',
        'measurement_session_number'
    ).order_by('pricelist_item__row_number', 'row_description')
    
    # گروه‌بندی بر اساس آیتم فهرست بها
    groups_dict = {}
    grand_total = Decimal('0.00')
    total_items = 0
    
    for item in session_items:
        pricelist_item = item.pricelist_item
        row_number = pricelist_item.row_number
        
        if row_number not in groups_dict:
            groups_dict[row_number] = {
                'row_number': row_number,
                'row_description': pricelist_item.description,
                'unit': pricelist_item.unit,
                'unit_price': item.unit_price or Decimal('0.00'),
                'items': [],
                'group_total': Decimal('0.00')
            }
        
        item_amount = item.get_total_item_amount() or Decimal('0.00')
        item_total = item.item_total or Decimal('0.00')
        
        # اطلاعات صورت جلسه برای لینک
        session = item.measurement_session_number
        session_info = {
            'id': session.id,
            'session_number': session.session_number,
            'session_date': session.session_date,
            'description': session.description,
            'url': reverse('sooratvaziat:session_detail', kwargs={'project_pk': project.pk, 'pk': session.id})
            }        
        groups_dict[row_number]['items'].append({
            'row_description': item.row_description,
            'length': item.length,
            'width': item.width,
            'height': item.height,
            'weight': item.weight,
            'count': item.count,
            'item_amount': item_amount,
            'item_total': item_total,
            'session': session_info,
            'unit_price': item.unit_price or Decimal('0.00')
        })
        
        groups_dict[row_number]['group_total'] += item_amount
        grand_total += item_amount
        total_items += 1
    
    # تبدیل دیکشنری به لیست
    groups = list(groups_dict.values())
    
    discipline_label = project.measurement_sessions.first().price_list.discipline
    
    context = {
        'title': f'ریز متره {discipline_label} - {project.project_name}',
        'project': project,
        'discipline_label': discipline_label,
        'groups': groups,
        'grand_total': grand_total,
        'total_items': total_items,
        'from_sessions_directly': True  # برای نمایش در تمپلیت
    }
    
    return render(request, 'sooratvaziat/riz_metre.html', context)

@login_required
def measurement_summary(request, pk, discipline):
    """
    نمایش خلاصه متره برای پروژه و رشته خاص - جدا برای هر فهرست بها
    """
    project = get_project_with_access(request.user, pk)
    
    # پیدا کردن تمام فهرست‌های بهای این رشته که در پروژه استفاده شده‌اند
    price_lists = PriceList.objects.filter(
        measurement_sessions__project=project,
        measurement_sessions__is_active=True,
        discipline_choice=discipline
    ).distinct()
    
    # اگر فهرست بها مشخصی از کوئری استرینگ ارسال شده
    price_list_id = request.GET.get('price_list_id')
    if price_list_id:
        selected_price_list = get_object_or_404(PriceList, id=price_list_id, discipline_choice=discipline)
    else:
        # انتخاب اولین فهرست بها به صورت پیش‌فرض
        selected_price_list = price_lists.first()
    
    summary_data = []
    grand_total = Decimal('0.00')
    total_items = 0
    
    if selected_price_list:
        # گروه‌بندی و جمع‌آوری داده‌ها فقط برای فهرست بها انتخاب شده
        summary_data = MeasurementSessionItem.objects.filter(
            measurement_session_number__project=project,
            measurement_session_number__is_active=True,
            measurement_session_number__price_list=selected_price_list,
            is_active=True
        ).select_related(
            'pricelist_item'
        ).values(
            'pricelist_item__row_number',
            'pricelist_item__description',
            'pricelist_item__unit',
            'pricelist_item__price'
        ).annotate(
            total_quantity=Sum('quantity'),
            total_amount=Sum('item_total'),
            sessions_count=Count('measurement_session_number', distinct=True),
            items_count=Count('id')
        ).order_by('pricelist_item__row_number')
        
        # تبدیل به ساختار داده‌ای مناسب
        summary_groups = []
        for item in summary_data:
            summary_groups.append({
                'row_number': item['pricelist_item__row_number'],
                'description': item['pricelist_item__description'],
                'unit': item['pricelist_item__unit'],
                'unit_price': Decimal(str(item['pricelist_item__price'])) if item['pricelist_item__price'] else Decimal('0.00'),
                'total_quantity': Decimal(str(item['total_quantity'])) if item['total_quantity'] else Decimal('0.00'),
                'total_amount': Decimal(str(item['total_amount'])) if item['total_amount'] else Decimal('0.00'),
                'sessions_count': item['sessions_count'],
                'items_count': item['items_count'],
                'formatted_unit_price': format_number_int(item['pricelist_item__price']) if item['pricelist_item__price'] else "۰",
                'formatted_total_quantity': format_number_int(item['total_quantity']) if item['total_quantity'] else "۰",
                'formatted_total_amount': format_number_int(item['total_amount']) if item['total_amount'] else "۰",
            })
            
            grand_total += Decimal(str(item['total_amount'])) if item['total_amount'] else Decimal('0.00')
            total_items += 1
    else:
        summary_groups = []

    context = {
        'title': f'خلاصه متره {selected_price_list.discipline if selected_price_list else "نامشخص"} - {project.project_name}',
        'project': project,
        'discipline_label': selected_price_list.discipline if selected_price_list else "نامشخص",
        'summary_groups': summary_groups,
        'grand_total': grand_total,
        'total_items': total_items,
        'formatted_grand_total': format_number_int(grand_total),
        'discipline': discipline,
        'price_lists': price_lists,
        'selected_price_list': selected_price_list,
    }
    
    return render(request, 'sooratvaziat/measurement_summary.html', context)

@login_required
def discipline_summary(request, pk):
    """
    نمایش خلاصه قیمت تمام رشته‌های پروژه
    """
    project = get_project_with_access(request.user, pk)
    
    # محاسبه مجموع هر رشته
    from django.db.models import Sum, Count
    
    discipline_data = MeasurementSessionItem.objects.filter(
        measurement_session_number__project=project,
        measurement_session_number__is_active=True,
        is_active=True
    ).values(
        'measurement_session_number__price_list__discipline_choice',
        'measurement_session_number__price_list__discipline'
    ).annotate(
        total_amount=Sum('item_total'),
        total_quantity=Sum('quantity'),
        sessions_count=Count('measurement_session_number', distinct=True),
        items_count=Count('id')
    ).order_by('measurement_session_number__price_list__discipline_choice')
    
    # تبدیل به ساختار داده‌ای مناسب
    disciplines_summary = []
    grand_total = Decimal('0.00')
    total_sessions = 0
    total_items = 0
    
    for item in discipline_data:
        discipline_choice = item['measurement_session_number__price_list__discipline_choice']
        discipline_name = item['measurement_session_number__price_list__discipline']
        total_amount = Decimal(str(item['total_amount'])) if item['total_amount'] else Decimal('0.00')
        
        # پیدا کردن label از choices
        discipline_label = dict(DisciplineChoices.choices).get(discipline_choice, discipline_name)
        
        disciplines_summary.append({
            'choice': discipline_choice,
            'name': discipline_name,
            'label': discipline_label,
            'total_amount': total_amount,
            'total_quantity': Decimal(str(item['total_quantity'])) if item['total_quantity'] else Decimal('0.00'),
            'sessions_count': item['sessions_count'],
            'items_count': item['items_count'],
            'formatted_amount': format_number_int(total_amount),
            'formatted_quantity': format_number_int(item['total_quantity']) if item['total_quantity'] else "۰",
        })
        
        grand_total += total_amount
        total_sessions += item['sessions_count']
        total_items += item['items_count']
    
    context = {
        'title': f'خلاصه رشته‌ها - {project.project_name}',
        'project': project,
        'disciplines_summary': disciplines_summary,
        'grand_total': grand_total,
        'total_sessions': total_sessions,
        'total_items': total_items,
        'formatted_grand_total': format_number_int(grand_total),
        'disciplines_count': len(disciplines_summary),
    }
    
    return render(request, 'sooratvaziat/discipline_summary.html', context)
    
@login_required
def session_list(request, pk):
    """
    لیست صورت جلسات یک پروژه
    """
    project = get_project_with_access(request.user, pk)

    # بررسی آیا کاربر دسترسی به این پروژه دارد
    has_access = (
        request.user.is_superuser or
        project.created_by == request.user or
        ProjectUser.objects.filter(
            project=project, 
            user=request.user, 
            is_active=True
        ).exists()
    )
    if not has_access:
        raise PermissionDenied("شما دسترسی به این پروژه را ندارید")
    try:
        # دریافت صورت جلسات مربوط به این پروژه
        sessions = MeasurementSession.objects.filter(
            project=project, 
            is_active=True
        ).annotate(
            items_count=Count('items', filter=Q(items__is_active=True))
        ).order_by('-created_at')

        # محاسبه آمار کلی
        total_sessions = sessions.count()
        approved_sessions = sessions.filter(status='approved').count()
        draft_sessions = sessions.filter(status='draft').count()

    except Exception as e:
        # اگر خطایی رخ داد، از مقادیر پیش‌فرض استفاده کن
        sessions = MeasurementSession.objects.filter(
            project=project, 
            is_active=True
        ).order_by('-created_at')
        
        # محاسبه دستی تعداد آیتم‌ها
        for session in sessions:
            session.items_count = session.items.filter(is_active=True).count()
            # مقدار پیش‌فرض برای status اگر وجود ندارد
            if not hasattr(session, 'status'):
                session.status = 'draft'
        
        total_sessions = sessions.count()
        approved_sessions = sessions.filter(status='approved').count() if hasattr(sessions.first(), 'status') else 0
        draft_sessions = sessions.filter(status='draft').count() if hasattr(sessions.first(), 'status') else total_sessions

    context = {
        'title': f'لیست صورت‌جلسات - {project.project_name}',
        'project': project,
        'sessions': sessions,
        'total_sessions': total_sessions,
        'approved_sessions': approved_sessions,
        'draft_sessions': draft_sessions,
    }
    return render(request, 'sooratvaziat/session_list.html', context)
    
@login_required
def session_create(request, project_pk):
    """
    ایجاد صورت جلسه جدید
    """
    project = get_project_with_access(request.user, project_pk)
    
    if request.method == 'POST':
        print("=" * 50)
        print("📥 دریافت داده‌های فرم ایجاد صورت جلسه")
        print("POST data:", dict(request.POST))
        print("=" * 50)
        
        form = MeasurementSessionForm(request.POST)
        
        # دیباگ: بررسی وضعیت فرم قبل از اعتبارسنجی
        print("🔍 وضعیت فرم قبل از is_valid():")
        print(f"   - discipline_filter value: {form.data.get('discipline_filter')}")
        print(f"   - price_list value: {form.data.get('price_list')}")
        print(f"   - price_list queryset count: {form.fields['price_list'].queryset.count()}")
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    session = form.save(commit=False)
                    session.project = project
                    session.created_by = request.user
                    session.modified_by = request.user
                    
                    # رشته از طریق فیلد discipline_filter انتخاب شده است
                    discipline_filter = form.cleaned_data.get('discipline_filter')
                    price_list = form.cleaned_data.get('price_list')
                    
                    print("✅ فرم معتبر است")
                    print(f"   - رشته انتخاب شده: {discipline_filter}")
                    print(f"   - فهرست بها انتخاب شده: {price_list} (ID: {price_list.id if price_list else 'None'})")
                    print(f"   - نام فهرست بها: {price_list.discipline if price_list else 'None'}")
                    print(f"   - شماره صورت جلسه: {form.cleaned_data.get('session_number')}")
                    print(f"   - وضعیت: {form.cleaned_data.get('status')}")
                    
                    session.save()
                    
                    print(f"✅ صورت جلسه با شماره {session.session_number} ایجاد شد")
                    print("=" * 50)
                    
                    messages.success(request, 'صورت جلسه با موفقیت ایجاد شد')
                    return redirect('sooratvaziat:session_detail', project_pk=project.pk, pk=session.pk)
                    
            except Exception as e:
                print(f"❌ خطا در ایجاد صورت جلسه: {e}")
                import traceback
                traceback.print_exc()
                print("=" * 50)
                messages.error(request, f'خطا در ایجاد صورت جلسه: {str(e)}')
        else:
            print("❌ فرم نامعتبر است")
            print("خطاهای فرم:", form.errors)
            print("داده‌های cleaned_data:", form.cleaned_data)
            print("=" * 50)
            messages.error(request, 'لطفا خطاهای فرم را برطرف کنید')
    else:
        # مقدار اولیه برای صورت جلسه جدید
        initial_data = {
            'session_date': timezone.now().date(),
            'status': 'draft'
        }
        form = MeasurementSessionForm(initial=initial_data)
        print("📝 فرم ایجاد جدید بارگذاری شد")
    
    context = {
        'title': 'ایجاد صورت جلسه جدید',
        'project': project,
        'form': form,
    }
    return render(request, 'sooratvaziat/session_form.html', context)

@login_required
def session_edit(request, project_pk, pk):
    """
    ویرایش صورت جلسه موجود
    """
    project = get_project_with_access(request.user, project_pk)
    
    session = get_object_or_404(
        MeasurementSession, 
        pk=pk, 
        project=project, 
        is_active=True
    )
    
    if request.method == 'POST':
        print("=" * 50)
        print("📥 دریافت داده‌های فرم ویرایش صورت جلسه")
        print("POST data:", dict(request.POST))
        print("=" * 50)
        
        form = MeasurementSessionForm(request.POST, instance=session)
        if form.is_valid():
            try:
                with transaction.atomic():
                    session = form.save(commit=False)
                    session.modified_by = request.user
                    
                    discipline_filter = form.cleaned_data.get('discipline_filter')
                    price_list = form.cleaned_data.get('price_list')
                    
                    print("✅ فرم معتبر است")
                    print(f"   - رشته انتخاب شده: {discipline_filter}")
                    print(f"   - فهرست بها انتخاب شده: {price_list} (ID: {price_list.id if price_list else 'None'})")
                    print(f"   - شماره صورت جلسه: {form.cleaned_data.get('session_number')}")
                    print(f"   - وضعیت: {form.cleaned_data.get('status')}")
                    
                    session.save()
                    
                    print(f"✅ صورت جلسه با شماره {session.session_number} ویرایش شد")
                    print("=" * 50)
                    
                    messages.success(request, 'صورت جلسه با موفقیت ویرایش شد')
                    return redirect('sooratvaziat:session_detail', project_pk=project.pk, pk=session.pk)
                    
            except Exception as e:
                print(f"❌ خطا در ویرایش صورت جلسه: {e}")
                print("=" * 50)
                messages.error(request, f'خطا در ویرایش صورت جلسه: {str(e)}')
        else:
            print("❌ فرم نامعتبر است")
            print("خطاهای فرم:", form.errors)
            print("=" * 50)
            messages.error(request, 'لطفا خطاهای فرم را برطرف کنید')
    else:
        form = MeasurementSessionForm(instance=session)
        print(f"📝 فرم ویرایش برای صورت جلسه {session.session_number} بارگذاری شد")
    
    context = {
        'title': f'ویرایش صورت جلسه - {session.session_number}',
        'project': project,
        'session': session,
        'form': form,
    }
    return render(request, 'sooratvaziat/session_form.html', context)

@login_required
def delete_session(request, project_pk, pk):
    """
    حذف نرم صورت جلسه
    """
    project = get_project_with_access(request.user, project_pk)
    
    session = get_object_or_404(
        MeasurementSession, 
        pk=pk, 
        project=project, 
        is_active=True
    )
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                session.is_active = False
                session.modified_by = request.user
                session.save()
                
                messages.success(request, 'صورت جلسه با موفقیت حذف شد')
                return redirect('sooratvaziat:session_list', pk=project.pk)
                
        except Exception as e:
            messages.error(request, f'خطا در حذف صورت جلسه: {str(e)}')
    
    return redirect('sooratvaziat:session_list', pk=project.pk)

@login_required
def MeasurementSessionView(request, pk):
    """
    نمایش صورت جلسه با گروه‌بندی آیتم‌ها
    """
    project = get_project_with_access(request.user, project_pk)
    
    try:
        # پیش‌فرض کردن آیتم‌ها
        item_queryset = MeasurementSessionItem.objects.filter(
            measurement_session_number__project=project,
            is_active=True
        ).select_related('pricelist_item').order_by('pricelist_item__row_number')

        sessions_qs = MeasurementSession.objects.filter(
            project=project,
            is_active=True
        ).prefetch_related(
            Prefetch('items', queryset=item_queryset)
        ).select_related('created_by').order_by('-session_date', '-created_at')

        sessions_data = []
        for session in sessions_qs:
            # ساخت گروه‌ها به صورت مرتب
            groups = OrderedDict()
            
            for item in session.items.all():
                pricelist_item = getattr(item, 'pricelist_item', None)
                if pricelist_item:
                    # استفاده از شماره ردیف برای گروه‌بندی
                    key = getattr(pricelist_item, 'row_number', None) or id(pricelist_item)
                    
                    if key not in groups:
                        groups[key] = {
                            'grouper': pricelist_item,
                            'items': [],
                            'group_total': Decimal('0.00')
                        }
                    
                    # محاسبه مبلغ آیتم
                    try:
                        item_amount = item.get_total_item_amount()
                        if not isinstance(item_amount, Decimal):
                            item_amount = Decimal(str(item_amount))
                        item_amount = item_amount.quantize(Decimal('1.00'), rounding=ROUND_HALF_UP)
                    except (ValueError, TypeError):
                        item_amount = Decimal('0.00')
                    
                    groups[key]['items'].append({
                        'instance': item,
                        'item_amount': item_amount,
                    })
                    groups[key]['group_total'] += item_amount

            # کمی کردن مجموع گروه‌ها
            for group in groups.values():
                group['group_total'] = group['group_total'].quantize(Decimal('1.00'), rounding=ROUND_HALF_UP)
                group['formatted_total'] = f"{group['group_total']:,.0f}"

            sessions_data.append({
                'instance': session,
                'groups': list(groups.values()),
                'session_total': sum(group['group_total'] for group in groups.values())
            })

    except Exception as e:
        messages.error(request, f"خطا در بارگذاری داده‌ها: {str(e)}")
        sessions_data = []

    context = {
        'sessions': sessions_data,
        'project': project,
    }
    return render(request, 'sooratvaziat/sooratjalase.html', context)

@login_required
def session_detail(request, project_pk, pk):
    """
    نمایش جزئیات صورت جلسه با قابلیت نمایش Revisionها
    """
    project = get_project_with_access(request.user, project_pk)
    session = get_object_or_404(MeasurementSession, pk=pk, project=project, is_active=True)

    # بررسی دسترسی به Revisionها
    user_can_view_revisions = can_view_revisions(request.user, project)
    user_can_edit_directly = can_edit_directly(request.user, project)
    
    # دریافت مستقیم آیتم‌ها
    active_items = session.items.filter(is_active=True).select_related('pricelist_item')
    
    # گروه‌بندی مستقیم در ویو
    grouped_items = []
    total_session_quantity = Decimal('0.00')  # محاسبه جمع کل صورت جلسه
    
    try:
        groups_dict = {}
        
        for item in active_items:
            if not item.pricelist_item:
                continue
                
            pl = item.pricelist_item
            key = f"{pl.row_number}_{pl.pk}"
            
            if key not in groups_dict:
                groups_dict[key] = {
                    'row_number': pl.row_number,
                    'description': pl.description,
                    'unit': pl.unit,
                    'total_quantity': Decimal('0.00'),  # جمع کل گروه
                    'notes': '',  # یادداشت گروه
                    'has_revisions': False,  # آیا Revision دارد
                }
            
            # محاسبه مقدار آیتم
            try:
                quantity = item.get_total_item_amount()
                if not isinstance(quantity, Decimal):
                    quantity = Decimal(str(quantity))
            except Exception as e:
                print(f"خطا در محاسبه مقدار آیتم {item.pk}: {e}")
                quantity = Decimal('0.00')
            
            # افزودن به جمع گروه
            groups_dict[key]['total_quantity'] += quantity
            
            # افزودن به جمع کل صورت جلسه
            total_session_quantity += quantity
            
            # بررسی وجود Revision
            if item.has_pending_revisions():
                groups_dict[key]['has_revisions'] = True
            
            # اگر یادداشت وجود دارد، اضافه کردن
            if item.notes and not groups_dict[key]['notes']:
                groups_dict[key]['notes'] = item.notes
        
        # تبدیل ساختار دیکشنری به لیست برای تمپلیت
        for key, group in groups_dict.items():
            formatted_group = {
                'row_number': group['row_number'],
                'description': group['description'],
                'unit': group['unit'],
                'total_quantity': group['total_quantity'].quantize(Decimal('1.00'), rounding=ROUND_HALF_UP),
                'notes': group['notes'],
                'has_revisions': group['has_revisions'],
            }
            
            grouped_items.append(formatted_group)
        
        # کمی کردن جمع کل
        total_session_quantity = total_session_quantity.quantize(Decimal('1.00'), rounding=ROUND_HALF_UP)
        
        print(f"✅ جمع کل صورت جلسه محاسبه شد: {total_session_quantity}")
        
    except Exception as e:
        print(f"❌ خطا در گروه‌بندی: {e}")
        import traceback
        traceback.print_exc()
        grouped_items = []
    
    # آمار کلی
    try:
        session_stats = session.get_session_stats()
    except Exception as e:
        session_stats = {
            'total_items': active_items.count(),
            'unique_pricelists': len(set(item.pricelist_item.pk for item in active_items if item.pricelist_item)),
        }
    
    # فرم‌های مدیریت آیتم‌ها
    item_form = MeasurementSessionItemForm(session=session)
    
    # لیست فهرست بها برای dropdown
    try:
        if session.price_list:
            pricelist_items = PriceListItem.objects.filter(
                price_list=session.price_list,
                is_active=True
            ).order_by('row_number')
        else:
            pricelist_items = PriceListItem.objects.none()
    except Exception as e:
        pricelist_items = PriceListItem.objects.none()
    
    context = {
        'title': f'جزئیات صورت جلسه - {getattr(session, "session_number", "بدون شماره")}',
        'project': project,
        'session': session,
        'grouped_items': grouped_items,
        'total_quantity': total_session_quantity,  # ارسال جمع کل به تمپلیت
        'session_stats': session_stats,
        'item_form': item_form,
        'pricelist_items': pricelist_items,
        'user_can_edit_directly': user_can_edit_directly,
        'user_can_view_revisions': user_can_view_revisions,
        'user_role': get_user_project_role(request.user, project),
    }
    
    return render(request, 'sooratvaziat/session_detail.html', context)

@login_required
def add_session_item(request, project_pk, session_pk):
    """
    افزودن آیتم جدید به صورت جلسه
    """
    project = get_project_with_access(request.user, project_pk)    
    
    session = get_object_or_404(
        MeasurementSession, 
        pk=session_pk, 
        project=project, 
        is_active=True
    )
    
    if request.method == 'POST':
        form = MeasurementSessionItemForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    item = form.save(commit=False)
                    item.measurement_session_number = session
                    item.created_by = request.user
                    item.modified_by = request.user
                    
                    # محاسبات خودکار
                    if item.pricelist_item and not item.unit_price:
                        item.unit_price = item._get_price_from_pricelist()
                    
                    item.quantity = item.get_total_item_amount()
                    item.item_total = item.quantity * item.unit_price
                    
                    item.save()
                    
                    # به‌روزرسانی تعداد آیتم‌های صورت جلسه
                    session.items_count = session.items.filter(is_active=True).count()
                    session.save(update_fields=['items_count'])
                    
                    messages.success(request, 'آیتم با موفقیت اضافه شد')
                    return redirect('sooratvaziat:session_detail', project_pk=project.pk, pk=session.pk)
                    
            except Exception as e:
                messages.error(request, f'خطا در ذخیره آیتم: {str(e)}')
        else:
            messages.error(request, 'لطفا خطاهای فرم را برطرف کنید')
    
    return redirect('sooratvaziat:session_detail', project_pk=project.pk, pk=session.pk)

@login_required
def edit_session_item(request, project_pk, session_pk, item_pk):
    """
    ویرایش آیتم صورت جلسه با قابلیت ثبت Revision
    """
    project = get_project_with_access(request.user, project_pk)
    session = get_object_or_404(MeasurementSession, pk=session_pk, project=project, is_active=True)
    item = get_object_or_404(MeasurementSessionItem, pk=item_pk, measurement_session_number=session, is_active=True)
    
    if request.method == 'POST':
        print("=" * 50)
        print("📥 دریافت داده‌های فرم ویرایش آیتم")
        print("POST data:", dict(request.POST))
        print("=" * 50)
        
        # استفاده از فرم ساده‌تر بدون instance اولیه
        form_data = request.POST.copy()
        
        try:
            with transaction.atomic():
                # بررسی نقش کاربر
                user_can_edit_directly = can_edit_directly(request.user, project)
                
                if user_can_edit_directly:
                    # کاربر پیمانکار یا سوپر یوزر - ویرایش مستقیم
                    # به‌روزرسانی فیلدها به صورت دستی
                    item.length = Decimal(form_data.get('length', 0) or 0)
                    item.width = Decimal(form_data.get('width', 0) or 0)
                    item.height = Decimal(form_data.get('height', 0) or 0)
                    item.count = Decimal(form_data.get('count', 1) or 1)
                    item.notes = form_data.get('notes', '')
                    item.modified_by = request.user
                    
                    # محاسبه مجدد مقدار
                    item.quantity = item.get_total_item_amount()
                    
                    # اگر unit_price وجود ندارد، از فهرست بها بگیر
                    if not item.unit_price and item.pricelist_item:
                        item.unit_price = item.pricelist_item.price
                    
                    item.item_total = item.quantity * item.unit_price
                    item.save()
                    
                    print(f"✅ آیتم {item.pk} با موفقیت ویرایش شد")
                    print(f"   - ابعاد جدید: {item.length} x {item.width} x {item.height}")
                    print(f"   - تعداد جدید: {item.count}")
                    print(f"   - مقدار جدید: {item.quantity}")
                    
                    messages.success(request, 'آیتم با موفقیت ویرایش شد')
                else:
                    # سایر کاربران - ایجاد Revision
                    revision_reason = form_data.get('revision_reason', 'ویرایش توسط کاربر')
                    
                    print(f"🔔 ایجاد Revision برای آیتم {item.pk}")
                    print(f"   - دلیل: {revision_reason}")
                    
                    revision = item.create_revision(
                        edited_by=request.user,
                        revision_reason=revision_reason,
                        new_length=Decimal(form_data.get('length', 0) or 0),
                        new_width=Decimal(form_data.get('width', 0) or 0),
                        new_height=Decimal(form_data.get('height', 0) or 0),
                        new_count=Decimal(form_data.get('count', 1) or 1),
                        new_notes=form_data.get('notes', '')
                    )
                    
                    # داده اصلی تغییر نمی‌کند
                    messages.success(request, 
                        'اصلاحیه با موفقیت ثبت شد. این تغییر پس از تأیید پیمانکار اعمال خواهد شد.'
                    )
                    
                    print(f"✅ Revision ایجاد شد: {revision.pk}")
                
                return redirect('sooratvaziat:session_detail', project_pk=project.pk, pk=session.pk)
                
        except Exception as e:
            print(f"❌ خطا در ویرایش آیتم: {e}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'خطا در ویرایش آیتم: {str(e)}')
    
    return redirect('sooratvaziat:session_detail', project_pk=project.pk, pk=session.pk)
    
@login_required
def delete_session_item(request, project_pk, session_pk, item_pk):
    """
    حذف نرم آیتم صورت جلسه
    """
    print(f"=== DELETE ITEM DEBUG ===")
    print(f"Project PK: {project_pk}, Session PK: {session_pk}, Item PK: {item_pk}")
    
    project = get_project_with_access(request.user, project_pk)    
    
    session = get_object_or_404(
        MeasurementSession, 
        pk=session_pk, 
        project=project, 
        is_active=True
    )
    
    item = get_object_or_404(
        MeasurementSessionItem, 
        pk=item_pk, 
        measurement_session_number=session,
        is_active=True
    )
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                print(f"Deleting item: {item.pk} - {item.row_description}")
                
                item.is_active = False
                item.modified_by = request.user
                item.save()
                
                # به‌روزرسانی تعداد آیتم‌های صورت جلسه
                session.items_count = session.items.filter(is_active=True).count()
                session.save(update_fields=['items_count'])
                
                print("Item deleted successfully")
                messages.success(request, 'آیتم با موفقیت حذف شد')
                
        except Exception as e:
            print(f"Error deleting item: {e}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'خطا در حذف آیتم: {str(e)}')
    
    return redirect('sooratvaziat:session_detail', project_pk=project.pk, pk=session.pk)

@login_required
def group_items_detail(request, project_pk, session_pk, pricelist_number):
    """
    نمایش جزئیات آیتم‌های یک گروه خاص
    """
    try:
        project = get_project_with_access(request.user, project_pk)
        session = get_object_or_404(MeasurementSession, pk=session_pk, project=project, is_active=True)
        
        print(f"🔍 درخواست جزئیات گروه - شماره فهرست: {pricelist_number}")
        print(f"   پروژه: {project_pk}, جلسه: {session_pk}")
        
        # دریافت آیتم‌های این گروه
        group_items = session.items.filter(
            is_active=True,
            pricelist_item__row_number=pricelist_number
        ).select_related('pricelist_item')
        
        print(f"   تعداد آیتم‌های یافت شده: {group_items.count()}")
        
        # محاسبه جمع‌های گروه
        total_quantity = sum(item.get_total_item_amount() for item in group_items)
        
        context = {
            'group_items': group_items,
            'pricelist_number': pricelist_number,
            'total_quantity': total_quantity,
        }
        
        return render(request, 'sooratvaziat/partials/group_items_detail.html', context)
        
    except Exception as e:
        print(f"❌ خطا در group_items_detail: {e}")
        import traceback
        traceback.print_exc()
        
        # بازگشت پاسخ خطا
        return render(request, 'sooratvaziat/partials/group_items_detail.html', {
            'error': f'خطا در بارگذاری جزئیات: {str(e)}'
        })
        
@login_required
def delete_session_items_by_pricelist(request, project_pk, session_pk, pricelist_number):
    """
    حذف تمام آیتم‌های یک گروه بر اساس شماره فهرست بها
    """
    project = get_project_with_access(request.user, project_pk)
    session = get_object_or_404(MeasurementSession, pk=session_pk, project=project, is_active=True)
    
    if request.method == 'POST':
        # حذف تمام آیتم‌های این گروه
        deleted_count, _ = session.items.filter(
            pricelist_item__row_number=pricelist_number,
            is_active=True
        ).update(is_active=False)
        
        messages.success(request, f'{deleted_count} آیتم از گروه {pricelist_number} حذف شد.')
        return redirect('sooratvaziat:session_detail', project_pk=project_pk, pk=session_pk)
    
    return redirect('sooratvaziat:session_detail', project_pk=project_pk, pk=session_pk)

# ***  revision  ***

@login_required
def get_item_revisions(request, item_pk):
    """
    دریافت تاریخچه تغییرات یک آیتم (AJAX)
    """
    project = get_project_with_access(request.user, project_pk)
    session = get_object_or_404(MeasurementSession, pk=session_pk, project=project, is_active=True)
    item = get_object_or_404(MeasurementSessionItem, pk=item_pk, measurement_session_number=session, is_active=True)
    
    # بررسی دسترسی
    if not can_view_revisions(request.user, project):
        return JsonResponse({'error': 'دسترسی غیرمجاز'}, status=403)
    
    revisions = item.get_active_revisions()
    
    html = render_to_string('sooratvaziat/partials/revisions_list.html', {
        'item': item,
        'revisions': revisions,
    })
    
    return JsonResponse({'html': html})

# ویو برای AJAX - دریافت فهرست‌های بها بر اساس رشته
@login_required
def get_price_lists_by_discipline(request):
    """
    دریافت فهرست‌های بها بر اساس رشته (AJAX)
    """
    discipline = request.GET.get('discipline')
    
    print(f"🔍 درخواست AJAX برای رشته: {discipline}")
    
    if discipline:
        try:
            price_lists = PriceList.objects.filter(
                discipline_choice=discipline,
                is_active=True
            ).values('id', 'discipline', 'year', 'discipline_choice')  # اضافه کردن discipline_choice
            
            price_lists_list = list(price_lists)
            print(f"✅ یافت شد {len(price_lists_list)} فهرست بها برای رشته {discipline}")
            
            # لاگ جزئیات فهرست‌های بها
            for pl in price_lists_list:
                print(f"   - ID: {pl['id']}, نام: {pl['discipline']}, سال: {pl['year']}, رشته: {pl['discipline_choice']}")
            
            return JsonResponse(price_lists_list, safe=False)
            
        except Exception as e:
            print(f"❌ خطا در دریافت فهرست‌های بها: {e}")
            return JsonResponse([], safe=False)
    
    print("⚠️ رشته مشخص نشده است")
    return JsonResponse([], safe=False)

# ویو برای AJAX - دریافت آیتم‌های فهرست بها
@login_required
def get_pricelist_items(request):
    """
    دریافت آیتم‌های یک فهرست بها (AJAX)
    """
    price_list_id = request.GET.get('price_list_id')
    
    if price_list_id:
        items = PriceListItem.objects.filter(
            price_list_id=price_list_id,
            is_active=True
        ).values('id', 'row_number', 'description', 'unit', 'price')
        
        items_list = list(items)
        return JsonResponse(items_list, safe=False)
    
    return JsonResponse([], safe=False)

@login_required
def project_financial_report_list(request):
    """
    لیست همه پروژه‌های قابل دسترسی کاربر برای مشاهده گزارش‌های مالی
    """
    # دریافت پروژه‌های قابل دسترسی کاربر
    if request.user.is_superuser:
        # سوپریوزر همه پروژه‌ها را می‌بیند
        projects = Project.objects.filter(is_active=True)
    else:
        # کاربران عادی فقط پروژه‌هایی که دسترسی دارند
        projects = Project.objects.filter(
            Q(is_active=True) & 
            (Q(created_by=request.user) | 
             Q(project_users__user=request.user, project_users__is_active=True))
        ).distinct()
    
    # جستجو
    search_query = request.GET.get('search', '').strip()
    if search_query:
        projects = projects.filter(
            Q(project_name__icontains=search_query) |
            Q(project_code__icontains=search_query) |
            Q(employer__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # مرتب‌سازی
    projects = projects.order_by('-created_at')
    
    # دریافت خلاصه‌های مالی از مدل ProjectFinancialSummary
    project_ids = projects.values_list('id', flat=True)
    financial_summaries = ProjectFinancialSummary.objects.filter(
        project_id__in=project_ids
    ).select_related('project')
    
    # ایجاد مپ برای دسترسی سریع
    financial_map = {summary.project_id: summary for summary in financial_summaries}
    
    # Pagination
    paginator = Paginator(projects, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # اضافه کردن اطلاعات مالی به پروژه‌ها
    for project in page_obj.object_list:
        financial_summary = financial_map.get(project.id)
        if financial_summary:
            project.financial_summary_data = {
                'total_amount': financial_summary.total_amount or Decimal('0.00'),
                'total_with_vat': financial_summary.total_with_vat or Decimal('0.00'),
                'progress_percentage': financial_summary.progress_percentage or Decimal('0.00'),
                'sessions_count': financial_summary.sessions_count or 0,
                'approved_sessions_count': financial_summary.approved_sessions_count or 0,
                'total_items_count': financial_summary.total_items_count or 0,
                'last_updated': financial_summary.last_updated,
                'formatted_total_amount': format_number_int(financial_summary.total_amount),
                'formatted_total_vat': format_number_int(financial_summary.total_with_vat),
                'progress_percentage_display': f"{financial_summary.progress_percentage:.1f}%" if financial_summary.progress_percentage else '۰%',
                'has_financial_data': bool(financial_summary.total_amount and financial_summary.total_amount > 0),
                'progress_class': _get_progress_class(financial_summary.progress_percentage or 0),
            }
        else:
            # اگر خلاصه مالی وجود ندارد، یک خلاصه خالی ایجاد کن
            project.financial_summary_data = get_empty_financial_summary()
    
    # آمار کلی
    total_projects = page_obj.paginator.count
    total_contract_amount = projects.aggregate(
        total=Sum('contract_amount')
    )['total'] or Decimal('0.00')
    
    total_measured_amount = sum(
        project.financial_summary_data['total_amount'] for project in page_obj.object_list
    )
    
    overall_progress = Decimal('0.00')
    if total_contract_amount > 0:
        overall_progress = (total_measured_amount / total_contract_amount) * 100
    
    context = {
        'projects': page_obj,
        'search_query': search_query,
        'total_projects': total_projects,
        'total_contract_amount': total_contract_amount,
        'formatted_total_contract': format_number_int(total_contract_amount),
        'total_measured_amount': total_measured_amount,
        'formatted_total_measured': format_number_int(total_measured_amount),
        'overall_progress_percentage': overall_progress,
        'formatted_overall_progress': f"{overall_progress:.1f}%",
        'title': 'گزارش‌های مالی پروژه‌ها',
        'page_title': 'لیست پروژه‌های مالی',
        'active_menu': 'financial_reports',
    }
    
    return render(request, 'sooratvaziat/project_financial_report_list.html', context)

def calculate_project_financial_summary(project):
    """
    محاسبه خلاصه مالی برای یک پروژه
    """
    try:
        # محاسبه مجموع مقادیر از صورت جلسات
        session_items = MeasurementSessionItem.objects.filter(
            measurement_session_number__project=project,
            is_active=True
        )
        
        total_amount = Decimal('0.00')
        total_with_vat = Decimal('0.00')
        sessions_count = MeasurementSession.objects.filter(
            project=project,
            is_active=True
        ).count()
        
        approved_sessions_count = MeasurementSession.objects.filter(
            project=project,
            is_active=True,
            status='approved'  # یا فیلد وضعیت تأیید
        ).count()
        
        total_items_count = session_items.count()
        
        # محاسبه مبالغ
        for item in session_items:
            try:
                qty = Decimal(str(item.get_total_item_amount() or 0))
                pl = item.pricelist_item
                unit_price = Decimal('0')
                for cand in ('price', 'unit_price', 'rate', 'baha'):
                    if hasattr(pl, cand):
                        val = getattr(pl, cand)
                        if val is not None:
                            unit_price = Decimal(str(val))
                            break
                item_total = qty * unit_price
                total_amount += item_total
                # اگر VAT داریم محاسبه شود
                total_with_vat += item_total  # اینجا می‌توان VAT را اضافه کرد
            except Exception as e:
                continue
        
        # محاسبه درصد پیشرفت
        progress_percentage = Decimal('0.00')
        if project.contract_amount and project.contract_amount > 0:
            progress_percentage = (total_amount / project.contract_amount) * 100
        
        return {
            'total_amount': total_amount.quantize(Decimal('1'), rounding=ROUND_HALF_UP),
            'total_with_vat': total_with_vat.quantize(Decimal('1'), rounding=ROUND_HALF_UP),
            'progress_percentage': progress_percentage.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
            'sessions_count': sessions_count,
            'approved_sessions_count': approved_sessions_count,
            'total_items_count': total_items_count,
            'last_updated': timezone.now(),
            'formatted_total_amount': format_number_int(total_amount),
            'formatted_total_vat': format_number_int(total_with_vat),
            'progress_percentage_display': f"{progress_percentage:.1f}%",
            'has_financial_data': total_amount > 0,
            'progress_class': _get_progress_class(progress_percentage),
        }
    
    except Exception as e:
        print(f"Error in calculate_project_financial_summary for project {project.id}: {e}")
        return get_empty_financial_summary()

def get_empty_financial_summary():
    """
    بازگرداندن یک خلاصه مالی خالی
    """
    return {
        'total_amount': Decimal('0.00'),
        'total_with_vat': Decimal('0.00'),
        'progress_percentage': Decimal('0.00'),
        'sessions_count': 0,
        'approved_sessions_count': 0,
        'total_items_count': 0,
        'last_updated': None,
        'formatted_total_amount': '۰',
        'formatted_total_vat': '۰',
        'progress_percentage_display': '۰%',
        'has_financial_data': False,
        'progress_class': 'danger',
    }

def _get_progress_class(percentage):
    """
    تعیین کلاس CSS بر اساس درصد پیشرفت
    """
    if percentage >= 80:
        return 'success'
    elif percentage >= 50:
        return 'warning'
    else:
        return 'danger'

@login_required
def project_financial_report(request, pk):
    """
    View برای گزارش مالی یک پروژه خاص
    """
    # بررسی دسترسی کاربر به پروژه
    project = get_project_with_access(request.user, pk)
    
    # محاسبه خلاصه مالی پروژه
    financial_summary = calculate_project_financial_summary(project)
    
    # دریافت آخرین صورت جلسات
    recent_sessions = MeasurementSession.objects.filter(
        project=project,
        is_active=True
    ).order_by('-created_at')[:10]
    
    # آمار پیشرفت بر اساس زمان
    progress_data = get_progress_timeline(project)
    
    context = {
        'project': project,
        'financial_summary': financial_summary,
        'recent_sessions': recent_sessions,
        'progress_data': progress_data,
        'title': f'گزارش مالی - {project.project_name}',
        'page_title': f'گزارش مالی پروژه {project.project_name}',
        'active_menu': 'financial_reports',
    }
    
    return render(request, 'sooratvaziat/project_financial_report.html', context)

def get_progress_timeline(project):
    """
    دریافت داده‌های پیشرفت زمانی پروژه
    """
    # این تابع می‌تواند داده‌های تاریخی پیشرفت را برگرداند
    # برای سادگی، یک ساختار نمونه برمی‌گردانیم
    return [
        {'date': '1403-01-01', 'progress': 0},
        {'date': '1403-02-01', 'progress': 10},
        {'date': '1403-03-01', 'progress': 25},
        # ...
    ]

@login_required
def session_financial_detail(request, session_id):
    """جزئیات مالی صورت‌جلسه - سریع"""
    session = get_object_or_404(
        MeasurementSession.objects.filter(
            project__in=Project.objects.filter(
                Q(created_by=request.user) | 
                Q(project_users__user=request.user, project_users__is_active=True)
            ).distinct()
        ), 
        id=session_id
    )
    
    # دریافت صورت وضعیت (بدون محاسبه!)
    financial_status = FinancialReportGenerator.get_session_financial_status(session_id)
    
    context = {
        'session': session,
        'financial_status': financial_status,
    }
    return render(request, 'sooratvaziat/session_financial_detail.html', context)

@login_required
def riz_mali_detail(request, pk, discipline_choice=None):
    """ریز مالی - سریع از دیتابیس"""
    project = get_project_with_access(request.user, project_pk)
    
    # دریافت ریز مالی (بدون محاسبه!)
    detailed_financials = FinancialReportGenerator.get_detailed_financial_report(
        pk, 
        discipline_choice
    )
    
    # خلاصه رشته
    discipline_summary = {}
    if discipline_choice:
        try:
            summary = ProjectFinancialSummary.objects.get(pk=pk)
            if discipline_choice == 'ab':
                discipline_summary = {
                    'quantity': summary.total_quantity_abnieh,
                    'amount': summary.total_amount_abnieh,
                }
            elif discipline_choice == 'mk':
                discipline_summary = {
                    'quantity': summary.total_quantity_mekanik,
                    'amount': summary.total_amount_mekanik,
                }
            elif discipline_choice == 'br':
                discipline_summary = {
                    'quantity': summary.total_quantity_bargh,
                    'amount': summary.total_amount_bargh,
                }
        except:
            pass
    
    context = {
        'project': project,
        'detailed_financials': detailed_financials,
        'discipline_summary': discipline_summary,
        'discipline_choice': discipline_choice,
        'discipline_label': dict(DisciplineChoices.choices).get(discipline_choice, ''),
    }
    return render(request, 'sooratvaziat/riz_mali_detail.html', context)

# ریز مالی پروژه
@login_required
def project_financial_report(request, pk):
    # فقط پروژه‌های کاربر جاری
    project = get_project_with_access(request.user, pk)
    
    disciplines_dict = {choice.value: choice.label for choice in DisciplineChoices}    
    
    # تبدیل تاریخ قرارداد به شمسی
    if project.contract_date:
        gregorian_date = project.contract_date
        jalali_date = jdatetime.date.fromgregorian(
            year=gregorian_date.year,
            month=gregorian_date.month,
            day=gregorian_date.day
        )
        contract_date_jalali = jalali_date.strftime("%Y/%m/%d")
    else:
        contract_date_jalali = "تعیین نشده"
    
    # استخراج رشته‌های منحصر به فرد از آیتم‌های موجود برای پروژه جاری
    disciplines_qs = MeasurementSessionItem.objects.filter(
        measurement_session_number__project=project,
        is_active=True
    ).values_list('pricelist_item__price_list__discipline_choice', flat=True).distinct()
    
    data_by_discipline = {}
    grand_total_quantity = Decimal('0')
    grand_total_amount = Decimal('0')
    total_items_count = 0
    
    for discipline in disciplines_qs:
        # فیلتر آیتم‌ها بر اساس رشته و پروژه جاری
        qs = MeasurementSessionItem.objects.filter(
            measurement_session_number__project=project,
            pricelist_item__price_list__discipline_choice=discipline,
            is_active=True
        ).select_related(
            'pricelist_item',
            'pricelist_item__price_list',
            'measurement_session_number'
        ).order_by('pricelist_item__row_number', 'id')
        
        rows = OrderedDict()
        for item in qs:
            pl = item.pricelist_item
            key = getattr(pl, 'row_number', None) or f"_id_{pl.pk}"
            if key not in rows:
                rows[key] = {
                    'pricelist_item': pl,
                    'row_number': getattr(pl, 'row_number', ''),
                    'unit': getattr(pl, 'unit', '') or '',
                    'total_qty': Decimal('0'),
                    'unit_price': Decimal('0'),
                    'line_total': Decimal('0'),
                }
            # تبدیل qty به Decimal
            qty = _to_decimal(item.get_total_item_amount(), places=0)
            rows[key]['total_qty'] += qty
        
        # محاسبه قیمت و جمع‌ها
        total_quantity = Decimal('0')
        total_amount = Decimal('0')
        items_count = len(rows)
        
        for r in rows.values():
            pl = r['pricelist_item']
            unit_price = Decimal('0')
            for cand in ('price', 'unit_price', 'rate', 'baha'):
                val = getattr(pl, cand, None)
                if val is not None:
                    unit_price = _to_decimal(val, places=0)
                    break
            r['unit_price'] = unit_price
            r['line_total'] = (r['total_qty'] * unit_price).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
            total_amount += r['line_total']
            total_quantity += r['total_qty']
            r['formatted_total_qty'] = format_number_int(r['total_qty'])
            r['formatted_unit_price'] = format_number_int(r['unit_price'])
            r['formatted_line_total'] = format_number_int(r['line_total'])
        
        # شماره‌گذاری فصل‌ها
        chapter_counters = defaultdict(int)
        numbered_rows = []
        prev_chapter = None
        for r in rows.values():
            rn = str(r['row_number'])
            chapter = rn[:2] if len(rn) >= 2 else "00"
            chapter_counters[chapter] += 1
            display_number = f"{chapter}-{chapter_counters[chapter]}"
            r['display_number'] = display_number
            r['chapter'] = chapter
            r['is_new_chapter'] = (chapter != prev_chapter)
            prev_chapter = chapter
            numbered_rows.append(r)
        
        if numbered_rows:
            # فقط صورت جلسات مربوط به این پروژه و رشته
            sessions = MeasurementSession.objects.filter(
                project=project,
                items__pricelist_item__price_list__discipline_choice=discipline,
                items__is_active=True
            ).distinct()
            
            data_by_discipline[discipline] = {
                'label': disciplines_dict.get(discipline, 'نامشخص'),
                'year': project.execution_year,
                'rows': numbered_rows,
                'total_quantity': total_quantity,
                'total_amount': total_amount,
                'items_count': items_count,
                'formatted_total_quantity': format_number_int(total_quantity),
                'formatted_total_amount': format_number_int(total_amount),
                'sessions': sessions,  # فقط صورت جلسات مرتبط با این پروژه و رشته
            }
            
            total_items_count += items_count
        
        grand_total_quantity += total_quantity
        grand_total_amount += total_amount
    
    grand_total_quantity = grand_total_quantity.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
    grand_total_amount = grand_total_amount.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
    
    # خروجی‌ها - اضافه کردن pk به پارامترها
    export = request.GET.get('export')
    if export == 'xlsx':
        return generate_excel_report(project, data_by_discipline, grand_total_quantity, grand_total_amount)
    elif export == 'pdf':
        return generate_pdf_report(request, project, data_by_discipline, grand_total_quantity, grand_total_amount)
    
    context = {
        'project': project,
        'contract_date_jalali': contract_date_jalali,
        'data_by_discipline': data_by_discipline,
        'grand_total_quantity': grand_total_quantity,
        'grand_total_amount': grand_total_amount,
        'total_items_count': total_items_count,
        'formatted_grand_total_quantity': format_number_int(grand_total_quantity),
        'formatted_grand_total_amount': format_number_int(grand_total_amount),
    }
    return render(request, 'sooratvaziat/project_financial_report.html', context)

# تابع برای تولید Excel
def generate_excel_report(project, data_by_discipline, grand_total_quantity, grand_total_amount):
    wb = Workbook()
    # شیت کلی
    ws_summary = wb.active
    ws_summary.title = "خلاصه پروژه"
    ws_summary.append(['پروژه', project.project_name])
    ws_summary.append(['کد پروژه', project.project_code])
    ws_summary.append([''])
    ws_summary.append(['دیسیپلین', 'سال', 'جمع مقدار', 'جمع مبلغ (ریال)'])
    for disc, data in data_by_discipline.items():
        ws_summary.append([data['label'], data['year'], data['total_quantity'], data['total_amount']])
    ws_summary.append(['جمع کل', '', grand_total_quantity, grand_total_amount])
    
    # استایل شیت خلاصه
    for row in ws_summary.iter_rows(min_row=1, max_row=ws_summary.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    
    # شیت برای هر دیسیپلین (ریز مالی)
    for disc, data in data_by_discipline.items():
        ws = wb.create_sheet(title=data['label'])
        headers = ['ردیف', 'شماره آیتم', 'شرح', 'واحد', 'مقدار', 'قیمت واحد', 'مبلغ کل']
        ws.append(headers)
        for r in data['rows']:
            ws.append([
                r['display_number'],
                r['row_number'],
                r['pricelist_item'].description,
                r['unit'],
                r['total_qty'],
                r['unit_price'],
                r['line_total'],
            ])
        ws.append(['', '', '', '', 'جمع', '', data['total_amount']])
        
        # استایل
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                if cell.row == 1:
                    cell.font = Font(bold=True)
    
    # ذخیره و بازگشت
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="financial_report_{project.project_code}.xlsx"'
    wb.save(response)
    return response

# تابع برای تولید PDF (با xhtml2pdf؛ HTML رو به PDF تبدیل می‌کنه)
def generate_pdf_report(request, project, data_by_discipline, grand_total_quantity, grand_total_amount):
    # رندر HTML اول (از تمپلیت مشابه)
    context = {
        'project': project,
        'data_by_discipline': data_by_discipline,
        'grand_total_quantity': grand_total_quantity,
        'grand_total_amount': grand_total_amount,
        'formatted_grand_total_quantity': format_number_int(grand_total_quantity),
        'formatted_grand_total_amount': format_number_int(grand_total_amount),
    }
  
    # اضافه کردن request به render_to_string برای دسترسی به request.user در template
    html = render_to_string('sooratvaziat/project_financial_report.html', context, request=request)
  
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="financial_report_{project.project_code}.pdf"'
  
    # تبدیل HTML به PDF با xhtml2pdf
    pisa_status = pisa.CreatePDF(
        html.encode('utf-8'),  # مطمئن شوید HTML به UTF-8 انکود شده برای پشتیبانی پارسی
        dest=response,
        encoding='utf-8'  # برای پشتیبانی از کاراکترهای پارسی
    )
  
    if pisa_status.err:
        # می‌تونید لاگ کنید: import logging; logger = logging.getLogger(__name__); logger.error("PDF generation error")
        return HttpResponse(f'خطا در تولید PDF: {pisa_status.err}', content_type='text/plain')
    
    return response

@login_required
def search(request):
    """
    View برای جستجو در پروژه‌ها و گزارش‌ها
    """
    query = request.GET.get('q', '').strip()
    search_results = []
    total_results = 0
    
    if query:
        # جستجو در پروژه‌ها
        projects = Project.objects.filter(
            Q(title__icontains=query) | 
            Q(description__icontains=query) |
            Q(user__name__icontains=query)
        ).distinct()[:10]
        
        # جستجو در گزارش‌ها
        reports = Report.objects.filter(
            Q(title__icontains=query) | 
            Q(content__icontains=query)
        ).distinct()[:10]
        
        # جستجو در ریز متره‌ها
        riz_metres = RizMetre.objects.filter(
            Q(title__icontains=query) | 
            Q(project__title__icontains=query)
        ).distinct()[:10]
        
        # ترکیب نتایج
        search_results = []
        
        # اضافه کردن پروژه‌ها
        for project in projects:
            search_results.append({
                'type': 'project',
                'title': project.title,
                'description': f"پروژه: {project.user.name if hasattr(project, 'user') else 'مشخص نشده'}",
                'url': reverse('sooratvaziat:project_detail', kwargs={'pk': project.pk}),
                'icon': 'bi-building',
                'highlight': query
            })
        
        # اضافه کردن گزارش‌ها
        for report in reports:
            search_results.append({
                'type': 'report',
                'title': report.title,
                'description': f"گزارش {report.report_type if hasattr(report, 'report_type') else ''}",
                'url': reverse('sooratvaziat:report_detail', kwargs={'report_id': report.id}),
                'icon': 'bi-file-earmark-text',
                'highlight': query
            })
        
        # اضافه کردن ریز متره‌ها
        for riz_metre in riz_metres:
            search_results.append({
                'type': 'riz_metre',
                'title': riz_metre.title,
                'description': f"ریز متره: {riz_metre.project.title if hasattr(riz_metre, 'project') else ''}",
                'url': reverse('sooratvaziat:riz_metre', kwargs={'id': riz_metre.project.id if hasattr(riz_metre, 'project') else 1}),
                'icon': 'bi-rulers',
                'highlight': query
            })
        
        total_results = len(search_results)
    
    # Pagination (اختیاری)
    page = request.GET.get('page', 1)
    paginator = Paginator(search_results, 10)
    page_obj = paginator.get_page(page)
    
    context = {
        'query': query,
        'search_results': page_obj,
        'total_results': total_results,
        'is_search_page': True,
        'page_obj': page_obj,
    }
    
    return render(request, 'sooratvaziat/search_results.html', context)

@require_http_methods(["GET", "POST"])
def search_ajax(request):
    """
    AJAX Search برای autocomplete
    """
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        query = request.GET.get('q', '').strip()
        results = []
        
        if len(query) >= 2:  # حداقل 2 کاراکتر
            # جستجوی سریع
            projects = Project.objects.filter(
                Q(title__istartswith=query) | 
                Q(title__icontains=query)
            )[:5]
            
            for project in projects:
                results.append({
                    'id': project.id,
                    'title': project.title,
                    'type': 'project',
                    'url': reverse('sooratvaziat:project_detail', kwargs={'pk': project.id}),
                    'icon': 'bi-building'
                })
            
            # جستجو در گزارش‌ها
            reports = Report.objects.filter(
                Q(title__istartswith=query)
            )[:5]
            
            for report in reports:
                results.append({
                    'id': report.id,
                    'title': report.title,
                    'type': 'report',
                    'url': reverse('sooratvaziat:report_detail', kwargs={'report_id': report.id}),
                    'icon': 'bi-file-earmark-text'
                })
        
        return JsonResponse({
            'results': results,
            'query': query,
            'count': len(results)
        })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

# اگر مدل‌ها وجود ندارند، view ساده بسازید
def search_simple(request):
    """
    View ساده برای جستجو در صورت عدم وجود مدل‌ها
    """
    query = request.GET.get('q', '').strip()
    
    # جستجوی ساده در دیتابیس (مثال)
    mock_results = []
    if query:
        # شبیه‌سازی نتایج
        mock_results = [
            {
                'title': f'پروژه {query}',
                'description': 'پروژه عمرانی یافت شده',
                'url': f'/projects/?search={query}',
                'icon': 'bi-building'
            },
            {
                'title': f'گزارش {query}',
                'description': 'گزارش مالی مرتبط',
                'url': f'/reports/?search={query}',
                'icon': 'bi-file-earmark-text'
            }
        ]
    
    context = {
        'query': query,
        'search_results': mock_results,
        'total_results': len(mock_results),
        'is_search_page': True,
    }
    
    return render(request, 'sooratvaziat/search_results.html', context)


    """
    دریافت کاربران یک پروژه (AJAX)
    """
    project = get_object_or_404(Project, pk=pk, is_active=True)
    
    if not project.has_access(request.user):
        return JsonResponse([], safe=False)
    
    users = project.project_users.filter(is_active=True).values(
        'user__id', 
        'user__username', 
        'user__first_name', 
        'user__last_name',
        'role__name'
    )
    
    users_list = list(users)
    return JsonResponse(users_list, safe=False)

